#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
批量调用 Dify API 处理脚本
从 Excel 文件读取问题，并行调用 Dify API，将结果输出到日志
"""

import argparse
import json
import logging
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Any, Optional

import requests
from openpyxl import load_workbook


# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class DifyAPIClient:
    """Dify API 客户端"""
    
    def __init__(self, api_url: str, jwt_token: str, jwt_chat: str, timeout: int = 400):
        """
        初始化客户端
        
        Args:
            api_url: Dify API 完整 URL
            jwt_token: JWT token（包含 Bearer 前缀）
            jwt_chat: 聊天JWT token
            timeout: 请求超时时间（秒）
        """
        self.api_url = api_url
        self.jwt_token = jwt_token
        self.jwt_chat = jwt_chat
        self.timeout = timeout
        self.headers = {
            'Content-Type': 'application/json',
            "Authorization": f"Bearer {self.jwt_token}"
        }

    def call_api(self, question_id: int, question_text: str) -> Dict[str, Any]:
        """
        调用 Dify API
        
        Args:
            question_id: 问题 ID
            question_text: 问题文本
            
        Returns:
            包含问题ID、问题文本和响应的字典
        """
        timestamp = int(time.time() * 1000)

        # 构造 query 内部对象
        query_obj = {
            "requestId": str(timestamp),
            "conversationId": "",
            "parentMessageId": "",
            "timezone": 8,
            "language": "CHINESE",
            "copilotAppCode": "DIAGNOSIS",
            "query": question_text,
            "metrics": None,
            "tables": None,
            "sheets": None,
            "metricModel": None,
            "queryArgs": {
                "defaultArgs": [],
                "placeholders": [],
                "default_filter": None
            },
            "alarm": None,
            "queryConfirm": None,
            "dashboardId": None,
            "historyId": None,
            "display": False,
            "classType": "input",
            "conversationName": question_text,
            "sqlFlag": 1
        }

        # 构造请求体
        payload = {
            "inputs": {
                "JWT": self.jwt_chat,
                "lang": "Chinese",
                "response_mode": "blocking",
                "sql_flag": 1
            },
            "query": json.dumps(query_obj, ensure_ascii=False),
            "response_mode": "blocking",
            "user": "yc"
        }

        try:
            logger.info(f"正在处理问题 ID {question_id}: {question_text[:50]}...")

            response = requests.post(
                self.api_url,
                headers=self.headers,
                json=payload,
                timeout=self.timeout
            )

            response.raise_for_status()
            response_data = response.json()

            logger.info(f"问题 ID {question_id} 处理成功")

            return {
                "id": question_id,
                "question": question_text,
                "response": response_data,
                "status": "success",
                "timestamp": datetime.now().isoformat()
            }

        except requests.exceptions.RequestException as e:
            logger.error(f"问题 ID {question_id} API 调用失败: {str(e)}")
            return {
                "id": question_id,
                "question": question_text,
                "response": None,
                "status": "error",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }
        except Exception as e:
            logger.error(f"问题 ID {question_id} 处理异常: {str(e)}")
            return {
                "id": question_id,
                "question": question_text,
                "response": None,
                "status": "error",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }


def read_questions_from_excel(excel_path: str) -> List[Dict[str, Any]]:
    """
    从 Excel 文件读取问题列表
    
    Args:
        excel_path: Excel 文件路径
        
    Returns:
        问题列表，每个元素包含 id 和 question
    """
    logger.info(f"开始读取 Excel 文件: {excel_path}")

    try:
        workbook = load_workbook(excel_path, read_only=True)
        sheet = workbook.active

        # 查找列名为 "qeruy" 的列
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # 查找 qeruy 列的索引
        qeruy_col_idx = None
        id_col_idx = None

        for idx, header in enumerate(header_row):
            if header and header.strip().lower() == 'qeruy':
                qeruy_col_idx = idx
            elif header and header.strip().upper() == 'ID':
                id_col_idx = idx

        if qeruy_col_idx is None:
            raise ValueError("未找到列名为 'qeruy' 的列")

        logger.info(f"找到 qeruy 列（索引: {qeruy_col_idx}）")
        if id_col_idx is not None:
            logger.info(f"找到 ID 列（索引: {id_col_idx}）")

        # 读取所有问题
        questions = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 获取问题文本
            if qeruy_col_idx < len(row):
                question_text = row[qeruy_col_idx]
            else:
                question_text = None

            # 跳过空行
            if not question_text or not str(question_text).strip():
                continue

            # 获取 ID
            if id_col_idx is not None and id_col_idx < len(row):
                question_id = row[id_col_idx]
                if question_id is None:
                    question_id = row_idx - 1
            else:
                question_id = row_idx - 1

            questions.append({
                "id": question_id,
                "question": str(question_text).strip()
            })

        workbook.close()

        logger.info(f"成功读取 {len(questions)} 个问题")
        return questions

    except Exception as e:
        logger.error(f"读取 Excel 文件失败: {str(e)}")
        raise


def process_questions_parallel(
    client: DifyAPIClient,
    questions: List[Dict[str, Any]],
    max_workers: int = 5
) -> List[Dict[str, Any]]:
    """
    并行处理问题列表
    
    Args:
        client: Dify API 客户端
        questions: 问题列表
        max_workers: 最大并发线程数
        
    Returns:
        结果列表
    """
    logger.info(f"开始并行处理 {len(questions)} 个问题，使用 {max_workers} 个线程")

    results = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        future_to_question = {
            executor.submit(client.call_api, q["id"], q["question"]): q
            for q in questions
        }

        # 收集结果
        for future in as_completed(future_to_question):
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                question = future_to_question[future]
                logger.error(f"任务执行异常 (ID: {question['id']}): {str(e)}")
                results.append({
                    "id": question["id"],
                    "question": question["question"],
                    "response": None,
                    "status": "error",
                    "error": str(e),
                    "timestamp": datetime.now().isoformat()
                })

    # 按 ID 排序结果
    results.sort(key=lambda x: x["id"])

    logger.info(f"所有任务完成，共处理 {len(results)} 个问题")
    return results


def extract_answer(response_data: Optional[Dict[str, Any]]) -> str:
    """
    从响应数据中提取 answer 字段
    
    Args:
        response_data: API 响应数据
        
    Returns:
        提取的答案文本
    """
    if not response_data:
        return "[错误: 无响应数据]"

    # 尝试提取 answer 字段
    answer = response_data.get("answer")
    if answer:
        return answer

    # 如果没有 answer 字段，返回提示
    return "[错误: 响应中未找到 answer 字段]"


def batch_process_questions(
    questions: List[str],
    api_url: str,
    jwt_token: str,
    jwt_chat: str,
    max_workers: int = 5,
    timeout: int = 400
) -> List[Dict[str, Any]]:
    """
    批量处理问题列表（可被外部调用）
    
    Args:
        questions: 问题文本列表
        api_url: Dify API URL
        jwt_token: JWT token
        jwt_chat: 聊天 JWT token
        max_workers: 并发线程数
        timeout: 超时时间（秒）
    
    Returns:
        结果列表
    """
    # 创建客户端
    client = DifyAPIClient(api_url, jwt_token, jwt_chat, timeout)
    
    # 构建问题字典（带ID）
    questions_with_id = [
        {"id": idx + 1, "question": q}
        for idx, q in enumerate(questions)
    ]
    
    # 调用原有的 process_questions_parallel
    return process_questions_parallel(client, questions_with_id, max_workers)


def save_results(
    results: List[Dict[str, Any]],
    output_dir: str = "output1"
):
    """
    输出结果统计信息到日志
    
    Args:
        results: 结果列表
        output_dir: 输出目录（保留参数以保持兼容性，但不再使用）
    """
    # 统计信息
    success_count = sum(1 for r in results if r['status'] == 'success')
    error_count = len(results) - success_count

    logger.info("="*80)
    logger.info(f"处理统计: 总计 {len(results)} 个问题, 成功 {success_count} 个, 失败 {error_count} 个")
    logger.info("="*80)


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='批量调用 Dify API')
    parser.add_argument('--jwt', default="app-ggdyQj8AEs4JaPER1TwWZXIr", help='JWT token（包含 Bearer 前缀）')
    parser.add_argument('--api-url', default="https://ai-aidq.sany.com.cn/v1/chat-messages", help='Dify API 完整 URL')
    parser.add_argument('--threads', type=int, default=5, help='并发线程数（默认: 5）')
    parser.add_argument('--input', default='query.xlsx', help='Excel 文件路径（默认: batch_query/query.xlsx）')
    parser.add_argument('--jwt_chat', default='Bearer eyJhbGciOiJIUzUxMiJ9.eyJvcmdhbml6YXRpb25JZCI6LTEsIm5hbWUiOiJhZG1pbiIsImlkIjotMSwiaXNBZG1pbiI6dHJ1ZSwidXNlcm5hbWUiOiJhZG1pbiIsInN1YiI6ImFkbWluIiwiaWF0IjoxNzYxNTc5NzExLCJleHAiOjg2NDAwMTc2MTU3OTcxMX0.rerteBn3_CX2UiRFVkDomOLzItCJqVqfWrRnPqwlw6Bui6hgGFy63cikOmNILM9r8urCFaTt4RQSsgiZ3ySWGw', help='chatbot的令牌')


    args = parser.parse_args()

    start_time = time.time()
    logger.info("="*80)
    logger.info("批量 Dify API 调用开始")
    logger.info(f"API URL: {args.api_url}")
    logger.info(f"输入文件: {args.input}")
    logger.info(f"并发线程数: {args.threads}")
    logger.info("="*80)

    try:
        # 读取问题
        questions = read_questions_from_excel(args.input)

        if not questions:
            logger.warning("未找到任何问题，程序退出")
            return

        # 创建客户端
        client = DifyAPIClient(args.api_url, args.jwt, args.jwt_chat)

        # 并行处理
        results = process_questions_parallel(client, questions, args.threads)

        # 保存结果
        save_results(results)

        elapsed_time = time.time() - start_time
        logger.info("="*80)
        logger.info(f"批量处理完成，总耗时: {elapsed_time:.2f} 秒")
        logger.info("="*80)

    except Exception as e:
        logger.error(f"程序执行失败: {str(e)}", exc_info=True)
        raise


if __name__ == "__main__":
    main()